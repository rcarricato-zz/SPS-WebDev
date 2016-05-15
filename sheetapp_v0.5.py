import openpyxl
import smtplib
import pprint
import os

from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.text import MIMEText

from werkzeug import secure_filename
from flask import Flask, abort, render_template, request, redirect, url_for

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '/tmp/sheetapp'
app.config['SMTP_HOST'] = '127.0.0.1'
app.config['SMTP_PORT'] = 25
app.config['SMTP_USER'] = '--------'
app.config['SMTP_PASS'] = '--------'
app.config['EMAIL_FROM'] = 'Provisioning Service <provisioning@realtimeconnect.com>'
app.config['EMAIL_SUBJECT'] = 'Your contact information'

attachments = [
	[ 'image001.jpg', 'jpeg' ],
	[ 'image002.png', 'png' ],
	[ 'image004.png', 'png' ],
	[ 'image006.png', 'png' ],
	[ 'image008.jpg', 'jpeg' ],
]

@app.route('/')
def upload():
	return render_template('upload.html', email=request.args.get('email', 'no'))

@app.route('/debug')
def debug():
	html = '<html><body><pre>' + pprint.pformat(app.config) + '</pre></body></html>'
	return html

def render_workbook(book):
	sheet = book.active
	domain1 = sheet['B2'].value
	domain2 = sheet['B3'].value
	phone = sheet['B4'].value
	users = []
	for row in sheet.iter_rows('A7:D%d' % (sheet.max_row + 1)):
		name = row[0].value
		if name:
			password = row[1].value
			confID = row[2].value
			email = row[3].value
			users.append({
				'name': name,
				'password': password,
				'confID': confID,
				'email': email,
			})
	return render_template('sheet.html', domain1=domain1, domain2=domain2, phone=phone, users=users)

@app.route('/test')
def test():
	book = openpyxl.load_workbook('Client1.xlsx')
	return render_workbook(book)

@app.route('/sheet/<filename>')
def sheet(filename):
	book = openpyxl.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))
	return render_workbook(book)

@app.route('/email/<filename>')
def email(filename):
	# Get the workbook open and grab company-wide info
	book = openpyxl.load_workbook(os.path.join(app.config['UPLOAD_FOLDER'], filename))
	sheet = book.active
	domain1 = sheet['B2'].value
	domain2 = sheet['B3'].value
	phone = sheet['B4'].value
	# Read in all the image attachments
	images = []
	for item in attachments:
		fp = open('static/' + item[0], 'rb')
		img = MIMEImage(fp.read(), item[1])
		fp.close()
		img.add_header('Content-ID', '<%s>' % item[0])
		images.append(img)
	# Connect to the mail server
	with smtplib.SMTP(app.config['SMTP_HOST'], app.config['SMTP_PORT']) as smtp:
		# Login
		smtp.ehlo()
		smtp.starttls()
		smtp.login(app.config['SMTP_USER'], app.config['SMTP_PASS'])
		# Loop over all rows of user info
		count = 0
		for row in sheet.iter_rows('A7:D%d' % (sheet.max_row + 1)):
			name = row[0].value
			if name:
				password = row[1].value
				confID = row[2].value
				email = row[3].value
				user = {
					'name': name,
					'password': password,
					'confID': confID,
					'email': email,
				}
				# Create the email text
				html = render_template('email.html', user=user, domain1=domain1, domain2=domain2, phone=phone)
				text = MIMEText(html, 'html')
				# Create the email message
				msg = MIMEMultipart('related')
				msg['To'] = email
				msg['From'] = app.config['EMAIL_FROM']
				msg['Subject'] = app.config['EMAIL_SUBJECT']
				msg.attach(text)
				for img in images:
					msg.attach(img)
				# Send the email
				smtp.sendmail(msg['From'], msg['To'], msg.as_string())
				print("Mail sent to %s" % email)
				count = count + 1
	return render_template('email_done.html', count=count)

@app.route('/submitSheet', methods=['POST'])
def submit():
	sheetUpload = request.files['sheetUpload']
	filename = secure_filename(sheetUpload.filename)
	dest = os.path.join(app.config['UPLOAD_FOLDER'], filename)
	if os.path.isdir(dest):
		abort(400);
	sheetUpload.save(dest)
	if request.form['email'] == 'yes':
		return redirect(url_for('email', filename=filename))
	return redirect(url_for('sheet', filename=filename))

if __name__ == '__main__':
	app.run(host='127.0.0.1', port=10001, debug=True)

